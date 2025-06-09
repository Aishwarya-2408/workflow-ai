import json
import os
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd

from utility import get_logger


class WorkflowTransformer:
    """
    Transforms potentially nested tree JSON data into MCW (Master Condition Workflow)
    and WCM (Workflow Condition Master) Excel outputs based on project-specific
    configuration settings.

    The class handles various input JSON structures by first flattening them into
    a consistent format representing paths to approval rule lists. It correctly
    processes keys containing special characters (like '.') and generates sequential
    Condition IDs based on a configurable starting point.

    Initialization requires a `project_id` to look up in the configuration file
    and the `project_config_path` pointing to the JSON configuration file.

    Example Usage:
        config = {
            "PROJECT_A": {
                "mcwId": "ALT_PROJ_A",
                "mcwTitle": "Project A Approval",
                "mcwProcess": "PROC_APPROVAL_A",
                "wcmStartConditionId": "WC1000",
                "wcmCurrency": "USD",
                "wcmDocument": "DOC_TYPE_A",
                "wcmConditionKeys": ["Region", "Department", "Category", "RiskLevel"]
            },
            "PROJECT_B": { ... }
        }
        with open("projects.json", "w") as f:
            json.dump(config, f)

        transformer = WorkflowTransformer("PROJECT_A", "projects.json")
        # Assuming input_data is your nested or flat JSON
        transformed_rules = transformer.transform_to_condition_rules(input_data)
        transformer.save_to_mcw_wcm(transformed_rules, "output_A_mcw.xlsx", "output_A_wcm.xlsx")
    """

    def __init__(self, project_id: str, project_config_path: str) -> None:
        """
        Initializes the WorkflowTransformer.

        Args:
            project_id: The identifier for the project whose configuration should be loaded.
            project_config_path: The file path to the JSON configuration file containing
                                 settings for one or more projects.

        Raises:
            FileNotFoundError: If the project_config_path does not exist.
            KeyError: If the project_id is not found within the configuration file.
            json.JSONDecodeError: If the configuration file is not valid JSON.
        """
        self.logger = get_logger() # Get logger instance
        self.project_id = project_id
        self.logger.info(f"Initializing WorkflowTransformer for project: '{project_id}'")

        if not os.path.isfile(project_config_path):
            self.logger.error(f"Configuration file not found at path: {project_config_path}")
            raise FileNotFoundError(f"Config file not found: {project_config_path}")

        try:
            with open(project_config_path, 'r', encoding='utf-8') as f:
                all_configs: Dict[str, Any] = json.load(f)
        except json.JSONDecodeError as e:
            self.logger.error(f"Failed to decode JSON from config file: {project_config_path} - {e}")
            raise e
        except Exception as e:
            self.logger.error(f"Error reading config file: {project_config_path} - {e}")
            raise e


        if project_id not in all_configs:
            self.logger.error(f"Project ID '{project_id}' not found in the configuration file: {project_config_path}")
            raise KeyError(f"Project ID '{project_id}' not found in config")

        self.config: Dict[str, Any] = all_configs[project_id]
        self.logger.info(f"Successfully loaded configuration for project '{project_id}'")

        # Helper function to safely get config values, supporting camelCase or snake_case keys
        def _get(*keys: str, default: Optional[Any] = None) -> Optional[Any]:
            for key in keys:
                if key in self.config:
                    return self.config[key]
            return default

        # --- Load Configuration Settings ---
        self.mcw_id: str = _get('mcw_id', 'mcwId', default='ALT0001')
        self.mcw_title: str = _get('mcw_title', 'mcwTitle', default='')
        self.mcw_process: str = _get('mcw_process', 'mcwProcess', default='')
        self.wcm_start_condition_id: str = _get('wcm_start_condition_id', 'wcmStartConditionId', default='WC001')
        self.wcm_currency: str = _get('wcm_currency', 'wcmCurrency', default='')
        self.wcm_document: str = _get('wcm_document', 'wcmDocument', default='')
        # Ensure condition_keys is always a list of strings
        loaded_keys = _get('wcm_condition_keys', 'wcmConditionKeys', default=[])
        self.condition_keys: List[str] = [str(k) for k in loaded_keys if isinstance(k, (str, int, float))]

        # --- Parse WCM Start Condition ID ---
        self.wcm_id_prefix: str = "WC"
        self.wcm_id_base_num: int = 0
        self.wcm_id_padding: int = 3 # Default padding

        if self.wcm_start_condition_id and isinstance(self.wcm_start_condition_id, str):
            num_part_str = ""
            prefix_part_str = self.wcm_start_condition_id # Default to full string if no digits found at end

            # Iterate backwards to extract trailing digits
            for i in range(len(self.wcm_start_condition_id) - 1, -1, -1):
                char = self.wcm_start_condition_id[i]
                if char.isdigit():
                    num_part_str = char + num_part_str
                else:
                    # Found the first non-digit, split point
                    prefix_part_str = self.wcm_start_condition_id[:i + 1]
                    break
            else:
                # If loop completed without break, the entire string might be digits or empty
                if num_part_str: # String is all digits
                    prefix_part_str = ""
                # Else: string was empty or no digits, keep defaults

            if num_part_str:
                try:
                    self.wcm_id_base_num = int(num_part_str)
                    # Padding is the length of the detected number part
                    self.wcm_id_padding = len(num_part_str)
                    # Ensure padding is at least 1 (e.g., for WC0)
                    self.wcm_id_padding = max(self.wcm_id_padding, 1)
                    self.wcm_id_prefix = prefix_part_str
                    self.logger.debug(f"Parsed start condition ID: prefix='{self.wcm_id_prefix}', base={self.wcm_id_base_num}, padding={self.wcm_id_padding}")
                except ValueError:
                    # Should not happen if isdigit() passed, but handle defensively
                    self.logger.warning(f"Could not parse numeric part '{num_part_str}' of wcm_start_condition_id: '{self.wcm_start_condition_id}'. Using defaults.")
                    # Revert to defaults
                    self.wcm_id_prefix = "WC"
                    self.wcm_id_base_num = 0
                    self.wcm_id_padding = 3
            else:
                self.logger.warning(f"Could not find numeric part at the end of wcm_start_condition_id: '{self.wcm_start_condition_id}'. Using defaults.")
                # Use default prefix, base, padding
                self.wcm_id_prefix = "WC"
                self.wcm_id_base_num = 0
                self.wcm_id_padding = 3
        else:
            self.logger.warning(f"Invalid or missing wcm_start_condition_id: '{self.wcm_start_condition_id}'. Using default prefix='WC', base=0, padding=3.")
            # Use default prefix, base, padding
            self.wcm_id_prefix = "WC"
            self.wcm_id_base_num = 0
            self.wcm_id_padding = 3


        # --- Log Final Settings ---
        self.logger.debug(f"MCW settings: id='{self.mcw_id}', title='{self.mcw_title}', process='{self.mcw_process}'")
        self.logger.debug(f"WCM settings: start_id='{self.wcm_start_condition_id}' (parsed as prefix='{self.wcm_id_prefix}', base={self.wcm_id_base_num}, padding={self.wcm_id_padding}), "
                          f"currency='{self.wcm_currency}', document='{self.wcm_document}'")
        self.logger.debug(f"Condition keys: {self.condition_keys}")


    def _flatten_nested_json(
        self,
        data: Any,
        parent_path: Optional[List[str]] = None,
        sep: str = "||" # Use a safe internal separator unlikely to be in keys
    ) -> List[Dict[str, Any]]:
        """
        Recursively flattens a nested dictionary structure down to the level
        where values are lists of approval rules.

        This helper method identifies leaf nodes which are lists, where each element
        is expected to be a dictionary containing 'label' and 'user' keys (or an empty list).
        Keys in the path containing dots or other special characters are preserved.

        Args:
            data: The dictionary or sub-dictionary to flatten.
            parent_path: The list of keys forming the path to the current data node.
            sep: The internal separator used for joining path keys during flattening.

        Returns:
            A list of dictionaries, where each dictionary represents a found leaf node
            and contains:
            - "path": A string representing the path to the leaf node, joined by `sep`.
                     (e.g., "Level1||Level2.Key||Level3")
            - "rules": The list of rule dictionaries found at that leaf node.
                     (e.g., [{'label': 'Role', 'user': 'Approver1'}, ...])
        """
        items: List[Dict[str, Any]] = []
        current_path: List[str] = parent_path or []

        if isinstance(data, dict):
            for k, v in data.items():
                # Ensure the key is treated as a string for the path
                current_key_str = str(k)
                new_path = current_path + [current_key_str]

                # --- Leaf Node Detection Logic ---
                # A leaf node is defined as a list where either:
                # 1. The list is empty.
                # 2. The list is not empty, and its first element is a dictionary
                #    containing both 'label' and 'user' keys. (Assumes homogeneity)
                is_rule_list = False
                if isinstance(v, list):
                    if not v: # Case 1: Empty list is a valid leaf
                        is_rule_list = True
                    # Case 2: Check structure of the first element if list is not empty
                    elif isinstance(v[0], dict) and 'label' in v[0] and 'user' in v[0]:
                         is_rule_list = True
                # --- End Leaf Node Detection ---

                if is_rule_list:
                    # Found a leaf node (the list of rules). Record its path and rules.
                    path_str = sep.join(new_path)
                    items.append({"path": path_str, "rules": v})
                    self.logger.debug(f"Flattened path found: {path_str}")
                elif isinstance(v, dict):
                    # It's a nested dictionary, so recurse deeper.
                    items.extend(self._flatten_nested_json(v, new_path, sep=sep))
                else:
                    # Value is neither a rule list nor a dictionary to recurse into.
                    # This path segment does not lead to a known rule structure. Log or ignore.
                    # path_str = sep.join(new_path) # For logging
                    # self.logger.debug(f"Ignoring non-dict/non-rule-list value at path: {path_str} (type: {type(v)})")
                    pass

        elif isinstance(data, list):
            # If the top-level input itself is a list, log a warning or try processing elements?
            # Current implementation expects a dict at the top.
            self.logger.warning(f"Input data at path '{sep.join(current_path)}' is a list, expected dictionary. Skipping list contents.")

        # Note: This function relies on the specific leaf node structure (list of dicts with 'label'/'user').
        # Adjust the `is_rule_list` logic if the structure of terminal rule lists changes.
        return items


    def transform_to_condition_rules(
        self,
        data: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Transforms the input JSON data (potentially nested or flat) into a list
        of condition rule dictionaries suitable for generating WCM and MCW outputs.

        It flattens the input using `_flatten_nested_json` to handle various structures,
        maps the flattened paths to configured condition keys, generates sequential
        Condition IDs, and constructs the condition string.

        Args:
            data: The input JSON data, which can be a nested dictionary or a flat
                  dictionary mapping keys directly to rule lists.

        Returns:
            A list of dictionaries, where each dictionary represents a condition rule row:
            {
                'Condition Id': str,  # e.g., "WC1001"
                'Condition': str,     # e.g., "Region=EMEA && Department=IT"
                'User Rule': list     # Original list of rule dicts, e.g., [{'label': 'Role', ...}]
            }
        """
        self.logger.info("Starting transformation to condition rules...")
        transformed: List[Dict[str, Any]] = []
        internal_sep = "||" # Separator used by _flatten_nested_json

        if not isinstance(data, dict) or not data:
             self.logger.warning("Input data is not a non-empty dictionary. Returning empty list.")
             return []

        # Step 1: Flatten the input data using the helper method
        flattened_items = self._flatten_nested_json(data, sep=internal_sep)

        # Step 2: Handle Fallback - If flattening failed but input looks flat
        if not flattened_items:
            self.logger.debug("Initial flattening yielded no results. Checking for flat structure fallback.")
            # Check if all top-level values are valid rule lists (using the same logic as _flatten_nested_json)
            is_flat_structure = all(
                isinstance(v, list) and
                (not v or (isinstance(v[0], dict) and 'label' in v[0] and 'user' in v[0]))
                for v in data.values()
            )
            if is_flat_structure:
                self.logger.info("Input data appears to be a flat structure. Processing directly.")
                # Convert the flat structure to the same format as flattened_items
                for key, rules_list in data.items():
                    # For a flat structure, the 'path' is just the top-level key itself
                    flattened_items.append({"path": str(key), "rules": rules_list})
            else:
                self.logger.warning("Flattening produced no items, and the input doesn't appear to be the expected flat structure. Check input data format. Returning empty list.")
                return [] # Return empty list if no processable items found

        # Step 3: Process the flattened items to create condition rules
        # Use the parsed base number, padding, and prefix from __init__
        base = self.wcm_id_base_num
        padding = self.wcm_id_padding
        prefix = self.wcm_id_prefix
        warned_about_depth = False # Flag to log depth warning only once

        for idx, item in enumerate(flattened_items):
            # Calculate the condition ID number, starting from base + 1
            # Ensures WC000 -> WC001, WC1000 -> WC1001 etc.
            counter = base + idx + 1
            full_path_str = item['path']
            user_rules = item['rules'] # This is the list of rule dicts [{label:.., user:..}, ..]

            # Split the flattened path string using the *internal* separator.
            # This correctly handles original keys that might contain dots.
            parts = full_path_str.split(internal_sep)

            # Map the path parts to the condition keys defined in the config
            cond_map: Dict[str, str] = {}
            if not self.condition_keys:
                # Case 1: No condition keys configured - use default names
                if idx == 0: # Log warning only for the first item
                    self.logger.warning("No 'wcm_condition_keys' defined in config. Using default names (Condition1, Condition2, ...).")
                for i, part in enumerate(parts):
                    key = f"Condition{i+1}"
                    cond_map[key] = part.strip() # Remove leading/trailing whitespace
            else:
                # Case 2: Use configured keys, falling back to defaults if path is deeper
                for i, part in enumerate(parts):
                    if i < len(self.condition_keys):
                        key = self.condition_keys[i]
                    else:
                        # Path depth exceeds configured keys - use default names for extra levels
                        key = f"Condition{i+1}"
                        if not warned_about_depth: # Log warning only once per transformation run
                            self.logger.warning(
                                f"Input data path ('{full_path_str}') has more levels ({len(parts)}) than "
                                f"configured condition keys ({len(self.condition_keys)}). "
                                f"Using default names (Condition{i+1}, ...) for extra levels."
                            )
                            warned_about_depth = True # Set flag
                    cond_map[key] = part.strip()

            # Construct the final condition string (e.g., "Key1=Val1 && Key2=Val2")
            # Filters out any potential empty key/value pairs just in case
            valid_conditions = {k: v for k, v in cond_map.items() if k and v}
            if not valid_conditions:
                self.logger.warning(f"Skipping condition for path '{full_path_str}' as it resulted in no valid key-value pairs.")
                continue # Skip this item if no valid conditions formed

            condition_str = ' && '.join(f"{k}={v}" for k, v in valid_conditions.items())

            # Format the Condition ID using prefix, counter, and dynamic padding
            try:
                cid = f"{prefix}{counter:0{padding}d}"
            except ValueError as e:
                self.logger.error(f"Error formatting Condition ID for counter {counter} with padding {padding}: {e}. Using default format.")
                cid = f"{prefix}{counter}" # Fallback format


            # Append the result dictionary for this path/rule set
            transformed.append({
                'Condition Id': cid,
                'Condition': condition_str,
                'User Rule': user_rules # Store the original list of rules
            })
            self.logger.debug(f"Generated {cid}: {condition_str}")

        self.logger.info(f"Successfully generated {len(transformed)} condition rules.")
        return transformed


    def save_to_mcw_wcm(
        self,
        transformed_data: List[Dict[str, Any]],
        mcw_output_path: str,
        wcm_output_path: str
    ) -> None:
        """
        Saves the transformed condition rules into two separate Excel files:
        MCW (Master Condition Workflow) and WCM (Workflow Condition Master).

        MCW file contains 'Basic Info' and 'Approval Path' sheets.
        WCM file contains a single sheet with condition details.

        Args:
            transformed_data: The list of condition rule dictionaries generated by
                              `transform_to_condition_rules`.
            mcw_output_path: The file path where the MCW Excel file will be saved.
            wcm_output_path: The file path where the WCM Excel file will be saved.
        """
        self.logger.info("Preparing data for MCW and WCM Excel files...")
        mcw_rows: List[Dict[str, Any]] = []
        wcm_rows: List[Dict[str, Any]] = []

        # --- Add Default Rejection Rule to MCW ---
        mcw_rows.append({
            'Id': self.mcw_id,
            'Condition Id': 'DEFAULT',
            'User Type': 'REJECT',
            'User Rule': 'Final_action:FINAL_REJECT'
        })

        # --- Prepare WCM Rows ---
        # Iterate through the transformed data to create rows for the WCM sheet
        for entry in transformed_data:
            cid = entry.get('Condition Id', 'ERROR_NO_ID')
            condition_str = entry.get('Condition', 'ERROR_NO_CONDITION')
            if cid == 'ERROR_NO_ID' or condition_str == 'ERROR_NO_CONDITION':
                 self.logger.error(f"Skipping entry due to missing Condition Id or Condition: {entry}")
                 continue

            wcm_rows.append({
                'Condition ID': cid,
                'Currency': self.wcm_currency,
                'Document': self.wcm_document,
                'Condition': condition_str,
                'Status': 'active', # Default status
                'Description': '' # Default description (can be enhanced later)
            })

            # --- Prepare MCW Rows (Approval Path) for the current condition ---
            # The 'User Rule' in entry is the list of approver dicts [{label:.., user:..}, ..]
            user_rule_list = entry.get('User Rule', [])
            if isinstance(user_rule_list, list):
                for rule_index, rule in enumerate(user_rule_list):
                    # Validate the rule structure
                    if not isinstance(rule, dict):
                        self.logger.warning(f"Skipping invalid item (expected dict) in 'User Rule' list for condition {cid}, index {rule_index}: {rule}")
                        continue

                    user = rule.get('user', 'N/A')
                    label = rule.get('label', 'N/A')

                    # Skip rules marked as N/A
                    if label == 'N/A' or user == 'N/A':
                        self.logger.debug(f"Skipping N/A rule for condition {cid}: label='{label}', user='{user}'")
                        continue

                    user_type = 'DYNAMIC' # Default for Role, may need adjustment for other labels
                    user_rule_formatted = ''

                    # Format the User Rule based on label and potentially specific users
                    if label == 'Role':
                        # Apply special logic for specific roles if needed
                        if user == 'Head of Department (OCN)- Cost Center Owner':
                            self.logger.debug(f"Applying special rule format for '{user}' on {cid}")
                            user_rule_formatted = (
                                'LABEL:Department_Approval_OC_Publish_CULT\n'
                                'LOOKUP_TABLE: CULT0019\n'
                                'RULE:OWNER_HIERARCHY_WITH_REQUIRED_APPROVAL_AUTHORITY\n'
                                'SKIP_INITIATOR: NO\n'
                                'APPROVAL_THRESHOLD:ANY'
                            )
                        else:
                            # Default format for Role label
                            user_rule_formatted = f"LABEL:{user} \nCRITERIA:ROLE={user} \nAPPROVAL_THRESHOLD:ANY"

                    # Add elif blocks here for other label types if needed
                    # elif label == 'User':
                    #     user_rule_formatted = f"LABEL:{user} \nCRITERIA:USER={user} \nAPPROVAL_THRESHOLD:ANY"
                    #     user_type = 'STATIC' # Example: Users might be static
                    # elif label == 'Group':
                    #     user_rule_formatted = f"LABEL:{user} \nCRITERIA:GROUP={user} \nAPPROVAL_THRESHOLD:ANY"
                    #     user_type = 'DYNAMIC' # Example: Groups are often dynamic

                    else:
                        self.logger.warning(f"Rule label '{label}' for condition {cid} is not explicitly handled. Skipping MCW row for this rule.")
                        continue # Skip adding this row if label is unknown

                    # Append the formatted row to MCW approval path data
                    mcw_rows.append({
                        'Id': self.mcw_id,
                        'Condition Id': cid,
                        'User Type': user_type,
                        'User Rule': user_rule_formatted
                    })
            else:
                 # Log error if 'User Rule' is not a list as expected
                 self.logger.error(f"Expected 'User Rule' to be a list for condition {cid}, but got {type(user_rule_list)}. Skipping MCW row generation for this condition.")

        # --- Build DataFrames ---
        # Basic Info for MCW
        mcw_basic_df = pd.DataFrame([{
            'Id': self.mcw_id,
            'Title': self.mcw_title,
            'Process': self.mcw_process,
            'Status': 'ACTIVE', # Default status
            'Description': '' # Default description
        }])

        # Approval Path for MCW
        mcw_approval_df = pd.DataFrame(mcw_rows)

        # WCM Data
        wcm_df = pd.DataFrame(wcm_rows)

        # --- Save MCW Excel File ---
        self.logger.info(f"Attempting to save MCW file to: {mcw_output_path}")
        try:
            with pd.ExcelWriter(mcw_output_path, engine='openpyxl') as writer:
                mcw_basic_df.to_excel(writer, sheet_name='Basic Info', index=False)
                sheet_names_to_format = ['Basic Info']

                # Only write and format Approval Path if there are rows (beyond the default)
                if len(mcw_approval_df) > 1: # Check if more than just the DEFAULT rule exists
                     mcw_approval_df.to_excel(writer, sheet_name='Approval Path', index=False)
                     sheet_names_to_format.append('Approval Path')
                     self.logger.debug(f"Writing {len(mcw_approval_df)} rows to 'Approval Path' sheet.")
                else:
                     # Still create the sheet but maybe empty or with just header if needed
                     pd.DataFrame(columns=mcw_approval_df.columns).to_excel(writer, sheet_name='Approval Path', index=False)
                     sheet_names_to_format.append('Approval Path') # Format header even if empty
                     self.logger.warning("No specific approval rules generated for MCW file. 'Approval Path' sheet contains only default rule or is empty.")

                self._format(writer, sheet_names_to_format)

            self.logger.info(f"Successfully saved MCW file: {mcw_output_path}")
        except PermissionError:
             self.logger.error(f"Permission denied saving MCW file: {mcw_output_path}. Check if the file is open or permissions are correct.", exc_info=False)
             raise
        except Exception as e:
             self.logger.error(f"An unexpected error occurred while saving MCW file '{mcw_output_path}': {e}", exc_info=True)
             raise # Re-raise the exception after logging

        # --- Save WCM Excel File ---
        self.logger.info(f"Attempting to save WCM file to: {wcm_output_path}")
        try:
             with pd.ExcelWriter(wcm_output_path, engine='openpyxl') as writer:
                 sheet_name = 'Sheet1' # Standard sheet name for WCM
                 # Only write and format if there are conditions
                 if not wcm_df.empty:
                     wcm_df.to_excel(writer, sheet_name=sheet_name, index=False)
                     self._format(writer, [sheet_name])
                     self.logger.debug(f"Writing {len(wcm_df)} rows to '{sheet_name}' sheet.")
                 else:
                      # Create an empty sheet with headers if no conditions were generated
                      pd.DataFrame(columns=wcm_df.columns).to_excel(writer, sheet_name=sheet_name, index=False)
                      self._format(writer, [sheet_name]) # Format the header of the empty sheet
                      self.logger.warning("No conditions generated for WCM file. '{sheet_name}' sheet is empty.")

             self.logger.info(f"Successfully saved WCM file: {wcm_output_path}")
        except PermissionError:
            self.logger.error(f"Permission denied saving WCM file: {wcm_output_path}. Check if the file is open or permissions are correct.", exc_info=False)
            raise
        except Exception as e:
            self.logger.error(f"An unexpected error occurred while saving WCM file '{wcm_output_path}': {e}", exc_info=True)
            raise # Re-raise the exception after logging


    def _format(self, writer: pd.ExcelWriter, sheet_names: List[str]) -> None:
        """
        Applies formatting (column width, header style, text wrap) to specified sheets
        within the Excel workbook managed by the pd.ExcelWriter.

        Args:
            writer: The pandas ExcelWriter object.
            sheet_names: A list of sheet names within the writer's workbook to format.
        """
        if not hasattr(writer, 'book'):
             self.logger.warning("ExcelWriter object does not have a 'book' attribute. Cannot apply formatting.")
             return

        workbook = writer.book
        # Define styles (consider making these configurable or constants)
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF") # White text
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Standard Blue fill
        header_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Default alignment for data cells (applied unless overridden)
        default_data_alignment = openpyxl.styles.Alignment(vertical='top', horizontal='left', wrap_text=False)
        # Alignment for cells that need wrapping (like User Rule)
        wrap_text_alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top', horizontal='left')

        self.logger.debug(f"Applying formatting to sheets: {sheet_names}")

        for sheet_name in sheet_names:
            if sheet_name not in writer.sheets:
                 self.logger.warning(f"Sheet '{sheet_name}' requested for formatting not found in the workbook.")
                 continue

            ws = writer.sheets[sheet_name]
            self.logger.debug(f"Formatting sheet: '{sheet_name}'")

            # Determine column index for 'User Rule' if this is the 'Approval Path' sheet
            user_rule_col_idx = None
            if sheet_name == 'Approval Path':
                 for col_idx, cell in enumerate(ws[1], 1): # Iterate header row (row 1)
                     if cell.value == 'User Rule':
                         user_rule_col_idx = col_idx
                         self.logger.debug(f"'User Rule' column found at index {user_rule_col_idx} in sheet '{sheet_name}'")
                         break

            # Iterate through columns to set width and apply header/data styles
            for col_idx, column_letter in enumerate(openpyxl.utils.get_column_letter(c_idx) for c_idx in range(1, ws.max_column + 1)):
                max_len = 0

                # Format header cell (Row 1)
                header_cell = ws.cell(row=1, column=col_idx + 1)
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = header_alignment
                if header_cell.value:
                    # Basic width calculation from header text
                    max_len = max(max_len, len(str(header_cell.value)))

                # Iterate through data cells (Row 2 onwards) in the current column
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    current_alignment = default_data_alignment # Start with default

                    # Check if this cell needs specific wrap-text formatting
                    is_user_rule_col = (sheet_name == 'Approval Path' and user_rule_col_idx is not None and (col_idx + 1) == user_rule_col_idx)

                    if cell.value:
                        cell_text = str(cell.value)
                        cell_len = 0
                        # If the cell contains newlines or is the User Rule column, prepare for wrapping
                        if '\n' in cell_text or is_user_rule_col:
                             current_alignment = wrap_text_alignment # Use wrap text alignment
                             # Calculate max width based on the longest line within the cell
                             lines = cell_text.split('\n')
                             cell_len = max(len(line) for line in lines) if lines else 0
                        else:
                             # Simple length for non-wrapping cells
                             cell_len = len(cell_text)

                        max_len = max(max_len, cell_len)

                    # Apply the determined alignment to the cell
                    cell.alignment = current_alignment

                # Set column width based on calculated max_len
                # Add buffer, apply min/max constraints
                adjusted_width = max_len + 3
                adjusted_width = min(adjusted_width, 60) # Max width cap
                adjusted_width = max(adjusted_width, 10) # Min width floor
                ws.column_dimensions[column_letter].width = adjusted_width
                # self.logger.debug(f"Set width for column {column_letter} in sheet '{sheet_name}' to {adjusted_width}")


# --- Example Usage Placeholder ---
if __name__ == "__main__":
    # This block is for demonstration or testing purposes.
    # Replace with your actual script logic.

    # 1. Set up logging (if not using utility)
    logger = get_logger()

    # 2. Define sample configuration (or load from file)
    sample_config = {
        "SIM_MALAYSIA": {
            "mcw_id": "ALT_SIM_MY",
            "mcw_title": "SIM Approval Workflow (Malaysia)",
            "mcw_process": "SIM_CREATE_SUPPLIER_MY",
            "wcm_start_condition_id": "WC00000", # Test start ID
            "wcm_currency": "MYR",
            "wcm_document": "SIM_SUPPLIER_DOC",
            "wcm_condition_keys": ["Subsidiary"]
        },
        "RFI_OCEANIA": {
            "mcwId": "ALT_RFI_OCN",
            "mcwTitle": "RFI Approval Oceania",
            "mcwProcess": "PROC_RFI_OCN",
            "wcmStartConditionId": "WC500",
            "wcmCurrency": "AUD",
            "wcmDocument": "RFI_DOC_OCN",
            "wcmConditionKeys": ["EventType", "Region", "Category", "RiskAssessment"]
        },
         "DEPT_HIERARCHY": {
            "mcw_id": "ALT_DEPT_HIER",
            "mcw_title": "Department Hierarchy Approval",
            "mcw_process": "PROC_DEPT_HIER",
            "wcm_start_condition_id": "WC100",
            "wcm_currency": "USD",
            "wcm_document": "DEPT_APPROVAL_DOC",
            "wcm_condition_keys": ["Division", "Department", "SubDepartment", "ChiefOfficer"]
        }
    }
    config_file = "temp_project_config.json"
    with open(config_file, "w", encoding='utf-8') as f:
        json.dump(sample_config, f, indent=4)
    logger.info(f"Created sample config file: {config_file}")


    # 3. Define sample input data for different structures
    flat_data_malaysia = {
      "AUTOMOTIVE CORPORATION (MALAYSIA) SDN. BHD.": [{"label": "Role","user": "ACM_CEO"}, {"label": "Role","user": "DRBHICOM_CPO"}],
      "DRB-HICOM BERHAD": [{"label": "Role", "user": "DRBHICOM_GCOO_CS"}, {"label": "Role", "user": "DRBHICOM_CPO"}],
      "MOTOSIKAL DAN ENJIN NASIONAL SDN. BHD.": [{"label": "Role", "user": "MODENAS_CEO"}, {"label": "Role", "user": "DRBHICOM_CPO"}],
    }

    nested_data_oceania = { # Shortened version of the RFI example
        "RFI": {
            "Oceania": {
                "IT": {
                    "Risk Assessment Results = APRA CPS234 (Information Security)": [
                        {"label": "Role", "user": "Oceania Requestor"},
                        {"label": "Role", "user": "Oceania ASO TISO"},
                        {"label": "Role", "user": "Procurement Head (SG)"},
                        {"label": "Role", "user": "Head of Department (OCN)- Cost Center Owner"}
                    ],
                    "Risk Assessment Results = NA": [
                        {"label": "Role", "user": "Oceania Requestor"},
                        {"label": "Role", "user": "Procurement Head (SG)"},
                        {"label": "Role", "user": "Head of Department (OCN)- Cost Center Owner"}
                    ]
                },
                 "Non- IT": {
                     "Risk Assessment Results = Both": [
                        {"label": "Role", "user": "Oceania Requestor"},
                        {"label": "Role", "user": "Oceania ASO TISO"},
                        {"label": "Role", "user": "Oceania Compliance Team"},
                        {"label": "Role", "user": "Procurement Head"},
                        {"label": "Role", "user": "Head of Department (OCN)- Cost Center Owner"}
                     ]
                 }
            }
        }
    }

    nested_data_dept = { # Shortened version of the Department Hierarchy example
        "Data, Tech & Delivery": {
            "Data & Insights": {
                "Analytics": {
                    "Chief Officer: Chief Technology & Data Officer": [
                        {"label": "Role", "user": "Head of Analytics & Research"},
                        {"label": "Role", "user": "General Manager, Data, Insights & Risk"},
                        {"label": "Role", "user": "Chief Technology & Data Officer"}
                    ],
                     "Head Of: Analytics & Research": [
                        {"label": "Role", "user": "Head of Analytics & Research"}
                     ]
                }
            },
            "Ent Architect & Tech": {
                 "Engineering": {
                      "Head Of: Engineering": [
                        {"label": "Role", "user": "Head of Engineering"}
                      ]
                 }
            }
        },
        "Finance & Investment Ops": {
            "Finance": {
                "Business Partnering": {
                     "General Manager: Finance Business Partnering": [
                        {"label": "Role", "user": "General Manager, Finance Business Partnering"},
                        {"label": "N/A", "user": "N/A"} # Test N/A rule skipping
                     ]
                 }
            }
        }
    }


    # 4. Create transformer instance and process each data structure
    try:
        logger.info("\n--- Processing Malaysia Data (Flat) ---")
        transformer_my = WorkflowTransformer("SIM_MALAYSIA", config_file)
        transformed_my = transformer_my.transform_to_condition_rules(flat_data_malaysia)
        if transformed_my:
             logger.info(f"Transformed Malaysia Data:\n{json.dumps(transformed_my, indent=2)}")
             transformer_my.save_to_mcw_wcm(transformed_my, "output_malaysia_mcw.xlsx", "output_malaysia_wcm.xlsx")
        else:
             logger.warning("Transformation yielded no results for Malaysia data.")

        logger.info("\n--- Processing Oceania Data (Nested) ---")
        transformer_ocn = WorkflowTransformer("RFI_OCEANIA", config_file)
        transformed_ocn = transformer_ocn.transform_to_condition_rules(nested_data_oceania)
        if transformed_ocn:
            logger.info(f"Transformed Oceania Data:\n{json.dumps(transformed_ocn, indent=2)}")
            transformer_ocn.save_to_mcw_wcm(transformed_ocn, "output_oceania_mcw.xlsx", "output_oceania_wcm.xlsx")
        else:
             logger.warning("Transformation yielded no results for Oceania data.")

        logger.info("\n--- Processing Department Data (Nested) ---")
        transformer_dept = WorkflowTransformer("DEPT_HIERARCHY", config_file)
        transformed_dept = transformer_dept.transform_to_condition_rules(nested_data_dept)
        if transformed_dept:
            logger.info(f"Transformed Department Data:\n{json.dumps(transformed_dept, indent=2)}")
            transformer_dept.save_to_mcw_wcm(transformed_dept, "output_dept_mcw.xlsx", "output_dept_wcm.xlsx")
        else:
            logger.warning("Transformation yielded no results for Department data.")


    except (FileNotFoundError, KeyError, json.JSONDecodeError) as e:
        logger.error(f"Initialization or processing failed: {e}")
    except Exception as e:
        logger.error(f"An unexpected error occurred during the example run: {e}", exc_info=True)
    finally:
        # Clean up the temporary config file
        if os.path.exists(config_file):
            try:
                os.remove(config_file)
                logger.info(f"Removed temporary config file: {config_file}")
            except OSError as e:
                logger.error(f"Error removing temporary config file {config_file}: {e}")